from openpyxl import load_workbook
import pandas as pd

# opening the source excel file
filename = 'C:\\Users\\deanejst\\Desktop\\sheet.xlsx'

# opening the destination excel file
filename1 = 'C:\\Users\\deanejst\\Desktop\\ae_open_wo.xlsx'

df = pd.read_excel(filename, sheet_name='Sheet1')
data = df.to_dict()

workbook = load_workbook(filename1)
worksheet = workbook.worksheets[0]

writer = pd.ExcelWriter(filename1, engine='openpyxl')

writer.book = load_workbook(filename1)
writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)

df.to_excel(writer, sheet_name='Sheet1', header=None, index=False,
            startcol=1, startrow=2)

writer.save()
